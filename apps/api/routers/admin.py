from __future__ import annotations

import asyncio
import csv
import io
import json
import re
from collections.abc import Iterable, Iterator, Sequence
from datetime import UTC, datetime, timedelta
from uuid import UUID, uuid4

import httpx
import openpyxl
import structlog
from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    Header,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from fastapi.responses import StreamingResponse
from sqlalchemy import String, case, cast, func, or_, select
from sqlalchemy.exc import IntegrityError

from apps.api.channels.voice import failover as voice_failover
from apps.api.channels.voice import plivo_client as voice_plivo
from apps.api.channels.whatsapp import client as wa_client
from apps.api.config import settings
from apps.api.core.llm import chat_completion
from apps.api.core.prompts import OUTBOUND_CALL_PROMPT, VOICE_APPEND, WHATSAPP_APPEND
from apps.api.core.tools import (
    TOOL_DEFINITIONS,
    _get_or_create_user_by_phone,
    _normalize_phone,
)
from apps.api.core.usage import get_credit_usage
from apps.api.db.models import (
    AccountUser,
    CallCampaign,
    CampaignTarget,
    Conversation,
    Lead,
    Message,
    Org,
    User,
)
from apps.api.deps import (
    AnalyticsScopeDep,
    CurrentUserDep,
    DbDep,
    RedisDep,
    RequestOrgDep,
    enforce_plan_limit,
    verify_admin_or_session,
)
from apps.api.rate_limit import limiter
from apps.api.redis_client import ERROR_COUNTER_KEY_FMT
from apps.api.schemas.admin import (
    CallingSettingsIn,
    CallingSettingsOut,
    KillSwitchIn,
    KillSwitchOut,
    OutboundCallIn,
    OutboundCallOut,
    OrgNumbersIn,
    OrgNumbersOut,
    OutboundWhatsappIn,
    OutboundWhatsappOut,
    PromptsOut,
    ScriptIn,
    ScriptOut,
    WhatsAppSettingsOut,
)
from apps.api.schemas.campaign import (
    CampaignCounts,
    CampaignCreateResult,
    CampaignDetailOut,
    CampaignOut,
    CampaignScheduleIn,
    CampaignStatusUpdateOut,
    CampaignTargetOut,
)
from apps.api.schemas.conversation import ConversationOut, ConversationSummaryOut, MessageOut
from apps.api.schemas.lead import (
    LEAD_QUALIFICATION_STATUSES,
    LEAD_STATUSES,
    LeadBulkImportIn,
    LeadDetailOut,
    LeadImportRow,
    LeadOut,
    LeadUpdateIn,
)
from apps.api.schemas.reports import ReportsCampaignRow, ReportsTimeseriesPoint

logger = structlog.get_logger(__name__)

router = APIRouter(
    prefix="/admin", tags=["admin"], dependencies=[Depends(verify_admin_or_session)]
)

# Redis keys / channels used by the control plane.
KILL_SWITCH_KEY = "veerox:kill_switch"
HUMAN_HANDOFF_QUEUE = "human_handoff_queue"


# Cost constants used by usd_spend_today. These mirror the planned values from
# implementation-plan.md §5.6 and intentionally live here rather than in
# apps/api/core/costs.py because that file is owned by worker 1 and may not
# exist yet. Once core/costs.py lands these can be re-imported from there.
_INPUT_USD_PER_TOKEN = 2.50 / 1_000_000  # $2.50 / 1M input tokens
_OUTPUT_USD_PER_TOKEN = 10.00 / 1_000_000  # $10.00 / 1M output tokens
_REALTIME_AUDIO_USD_PER_SECOND = 0.30 / 60.0  # $0.30 / minute of realtime audio


def _today_start_utc() -> datetime:
    return datetime.now(UTC).replace(hour=0, minute=0, second=0, microsecond=0)


@router.get("/stats")
async def get_stats(
    db: DbDep,
    redis: RedisDep,
    scope_org_id: AnalyticsScopeDep,
    x_admin_token: str | None = Header(None),
) -> dict:
    """Today's activity counters. Scoped to the caller's own org for a
    normal customer, platform-wide only for the platform admin — see
    `resolve_analytics_scope_org_id`, where `scope_org_id is None` is what
    "count every org" means.

    NOT cached, deliberately: the dashboard explicitly invalidates and
    refetches this query right after actions like placing a call or sending
    WhatsApp (useOutbound.ts) specifically to show the just-updated count —
    a server-side cache here would silently defeat that invalidation and
    make the dashboard look broken ("I did the thing but the number didn't
    change") for the length of the cache window. A round trip is worth
    paying for correctness here.
    """
    today_start = _today_start_utc()

    def scoped(stmt, model):  # noqa: ANN001, ANN202 — SQLAlchemy Select generics
        return stmt if scope_org_id is None else stmt.where(model.org_id == scope_org_id)

    def _count_if(condition):  # noqa: ANN001, ANN202 — SQLAlchemy boolean expr
        return func.sum(case((condition, 1), else_=0))

    def _scalar(stmt, model):  # noqa: ANN001, ANN202 — SQLAlchemy Select generics
        return scoped(stmt, model).scalar_subquery()

    # All 9 counters folded into ONE round trip via scalar subqueries, rather
    # than 4 separate awaited queries (one per table) — DB is a remote Neon
    # instance (see .env), so every extra round trip is real latency on a
    # page loaded on every dashboard visit. Postgres plans/executes these
    # subqueries together; only the network round trip is what we're saving.
    combined_stmt = select(
        _scalar(
            select(func.count()).select_from(User).where(User.created_at >= today_start), User
        ).label("users_today"),
        _scalar(
            select(func.count())
            .select_from(Conversation)
            .where(Conversation.channel == "voice", Conversation.started_at >= today_start),
            Conversation,
        ).label("calls_today"),
        _scalar(
            select(func.coalesce(_count_if(Message.channel == "whatsapp"), 0)).where(
                Message.created_at >= today_start
            ),
            Message,
        ).label("whatsapp_messages_today"),
        _scalar(
            select(func.coalesce(func.sum(Message.tokens_in), 0)).where(
                Message.created_at >= today_start
            ),
            Message,
        ).label("tokens_in_sum"),
        _scalar(
            select(func.coalesce(func.sum(Message.tokens_out), 0)).where(
                Message.created_at >= today_start
            ),
            Message,
        ).label("tokens_out_sum"),
        _scalar(
            select(func.coalesce(func.sum(Message.audio_secs), 0.0)).where(
                Message.created_at >= today_start
            ),
            Message,
        ).label("audio_secs_sum"),
        _scalar(
            select(func.count()).select_from(Lead).where(Lead.created_at >= today_start), Lead
        ).label("leads_today"),
        _scalar(
            select(func.coalesce(_count_if(Lead.channel == "voice"), 0)).where(
                Lead.created_at >= today_start
            ),
            Lead,
        ).label("leads_today_voice"),
        _scalar(
            select(func.coalesce(_count_if(Lead.channel == "whatsapp"), 0)).where(
                Lead.created_at >= today_start
            ),
            Lead,
        ).label("leads_today_whatsapp"),
    )

    # error_count_today — Redis counter keyed by today's UTC date. Written by
    # apps.api.redis_client.record_error(), called from each channel/worker's
    # top-level catch-all (whatsapp adapter, voice realtime bridge, campaign
    # dialer tick). It's a single platform-wide counter with no org
    # dimension, so it's only meaningful — and only reported — for the
    # platform admin; a customer org would otherwise see error counts driven
    # by other tenants' traffic. Runs concurrently with the DB round trip
    # above — separate connection, nothing shared to serialize on.
    async def _load_error_count() -> int:
        if scope_org_id is not None:
            return 0
        today_key = ERROR_COUNTER_KEY_FMT.format(date=datetime.now(UTC).date().isoformat())
        raw_err = await redis.get(today_key)
        try:
            return int(raw_err) if raw_err is not None else 0
        except (TypeError, ValueError):
            return 0

    combined_result, error_count_today = await asyncio.gather(
        db.execute(combined_stmt), _load_error_count()
    )
    (
        users_today,
        calls_today,
        whatsapp_messages_today,
        tokens_in_sum,
        tokens_out_sum,
        audio_secs_sum,
        leads_today,
        leads_today_voice,
        leads_today_whatsapp,
    ) = combined_result.one()

    usd_spend_today = (
        float(tokens_in_sum) * _INPUT_USD_PER_TOKEN
        + float(tokens_out_sum) * _OUTPUT_USD_PER_TOKEN
        + float(audio_secs_sum) * _REALTIME_AUDIO_USD_PER_SECOND
    )

    return {
        "users_today": users_today,
        "calls_today": calls_today,
        "whatsapp_messages_today": whatsapp_messages_today,
        "leads_today": leads_today,
        "leads_today_voice": leads_today_voice,
        "leads_today_whatsapp": leads_today_whatsapp,
        "p50_turn_latency_ms": None,
        "usd_spend_today": round(usd_spend_today, 6),
        "error_count_today": error_count_today,
    }


@router.get("/reports/timeseries", response_model=list[ReportsTimeseriesPoint])
async def get_reports_timeseries(
    db: DbDep,
    scope_org_id: AnalyticsScopeDep,
    x_admin_token: str | None = Header(None),
    days: int = Query(30, ge=1, le=365),
) -> list[ReportsTimeseriesPoint]:
    """Daily trend data for the sales reports page — the historical
    counterpart to GET /admin/stats, which only ever covers "today". Carries
    the same org scoping: a customer sees only their own org's trend, the
    platform admin sees every org's combined.

    Buckets by ``func.date(...)`` rather than ``date_trunc`` so the same
    query works against both Postgres (production) and SQLite (tests).
    """
    since = datetime.now(UTC) - timedelta(days=days)

    def scoped(stmt, model):  # noqa: ANN001, ANN202 — SQLAlchemy Select generics
        return stmt if scope_org_id is None else stmt.where(model.org_id == scope_org_id)

    # Keys are cast to str immediately — func.date(...) returns a raw
    # `datetime.date` on Postgres, but `all_days`/the lookups below are
    # strings, so leaving these as date-object keys means every .get(day)
    # below silently misses and falls back to 0 (calls/whatsapp_messages
    # always reading 0 on the reports chart even with real activity).
    calls_by_day = {
        str(day_val): count_val
        for day_val, count_val in (
            await db.execute(
                scoped(
                    select(func.date(Conversation.started_at), func.count()).where(
                        Conversation.channel == "voice", Conversation.started_at >= since
                    ),
                    Conversation,
                ).group_by(func.date(Conversation.started_at))
            )
        ).all()
    }
    # One round trip per table instead of one per counter (conditional SUM in
    # place of a separate query for each breakdown) — DB is a remote Neon
    # instance (see .env), so every extra query here is real page-load
    # latency on the Reports page. NOTE: these 3 queries can't safely run
    # concurrently via asyncio.gather — they'd need to share this request's
    # single `db` session, which isn't safe for concurrent statements, and
    # opening separate AsyncSessionLocal() sessions here bypasses the test
    # suite's DB fixture override (broke test_reports_endpoints.py — reverted).
    def _count_if(condition):  # noqa: ANN001, ANN202 — SQLAlchemy boolean expr
        return func.sum(case((condition, 1), else_=0))

    message_rows = (
        await db.execute(
            scoped(
                select(
                    func.date(Message.created_at),
                    func.coalesce(_count_if(Message.channel == "whatsapp"), 0),
                    func.coalesce(func.sum(Message.tokens_in), 0),
                    func.coalesce(func.sum(Message.tokens_out), 0),
                    func.coalesce(func.sum(Message.audio_secs), 0.0),
                ).where(Message.created_at >= since),
                Message,
            ).group_by(func.date(Message.created_at))
        )
    ).all()
    whatsapp_by_day: dict[str, int] = {}
    usd_spend_by_day: dict[str, float] = {}
    for day_val, whatsapp_count, tokens_in_sum, tokens_out_sum, audio_secs_sum in message_rows:
        day = str(day_val)
        whatsapp_by_day[day] = whatsapp_count
        usd_spend_by_day[day] = (
            float(tokens_in_sum) * _INPUT_USD_PER_TOKEN
            + float(tokens_out_sum) * _OUTPUT_USD_PER_TOKEN
            + float(audio_secs_sum) * _REALTIME_AUDIO_USD_PER_SECOND
        )

    lead_rows = (
        await db.execute(
            scoped(
                select(
                    func.date(Lead.created_at),
                    Lead.channel,
                    func.count(),
                    func.coalesce(_count_if(Lead.status == "qualified"), 0),
                ).where(Lead.created_at >= since),
                Lead,
            ).group_by(func.date(Lead.created_at), Lead.channel)
        )
    ).all()
    leads_voice_by_day: dict[str, int] = {}
    leads_whatsapp_by_day: dict[str, int] = {}
    qualified_by_day: dict[str, int] = {}
    for day_val, channel_val, count_val, qualified_val in lead_rows:
        day = str(day_val)
        if channel_val in ("voice", "whatsapp"):
            target = leads_voice_by_day if channel_val == "voice" else leads_whatsapp_by_day
            target[day] = count_val
        qualified_by_day[day] = qualified_by_day.get(day, 0) + qualified_val

    all_days = sorted(
        {str(d) for d in calls_by_day}
        | {str(d) for d in whatsapp_by_day}
        | set(leads_voice_by_day)
        | set(leads_whatsapp_by_day)
        | set(qualified_by_day)
        | set(usd_spend_by_day)
    )
    return [
        ReportsTimeseriesPoint(
            date=day,
            calls=calls_by_day.get(day, 0),
            whatsapp_messages=whatsapp_by_day.get(day, 0),
            leads_voice=leads_voice_by_day.get(day, 0),
            leads_whatsapp=leads_whatsapp_by_day.get(day, 0),
            qualified_count=qualified_by_day.get(day, 0),
            usd_spend=round(usd_spend_by_day.get(day, 0.0), 6),
        )
        for day in all_days
    ]


@router.get("/reports/campaigns", response_model=list[ReportsCampaignRow])
async def get_reports_campaigns(
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> list[ReportsCampaignRow]:
    """Per-campaign conversion table for the reports page — reuses
    ``_campaign_counts`` (defined further down) so this stays consistent
    with the counts already shown on the campaigns list."""
    org_id = org
    stmt = (
        select(CallCampaign)
        .where(CallCampaign.org_id == org_id)
        .order_by(CallCampaign.created_at.desc())
    )
    campaigns = (await db.execute(stmt)).scalars().all()
    counts_by_id = await _campaign_counts_bulk(db, [c.id for c in campaigns])

    rows: list[ReportsCampaignRow] = []
    for campaign in campaigns:
        counts = counts_by_id[campaign.id]
        qualification_rate = (
            counts.qualified / counts.completed if counts.completed else None
        )
        rows.append(
            ReportsCampaignRow(
                id=campaign.id,
                name=campaign.name,
                channel=campaign.channel,
                status=campaign.status,
                counts=counts,
                qualification_rate=qualification_rate,
            )
        )
    return rows


@router.get("/reports/export.xlsx")
async def export_reports_xlsx(
    db: DbDep,
    scope_org_id: AnalyticsScopeDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
    days: int = Query(30, ge=1, le=365),
) -> StreamingResponse:
    """The Reports page's "Download report" button — a workbook with the
    same two tables the page shows (daily trend, campaign conversion),
    scoped to the same `days` window, as a sales-ops-friendly alternative to
    reading the dashboard on screen. Reuses the JSON endpoints' own query
    logic directly rather than duplicating it."""
    timeseries = await get_reports_timeseries(db, scope_org_id, x_admin_token, days)
    campaign_rows = await get_reports_campaigns(db, org, x_admin_token)

    workbook = openpyxl.Workbook()
    trend_sheet = workbook.active
    trend_sheet.title = "Daily trend"
    trend_sheet.append(["Date", "Calls", "WhatsApp", "Qualified", "Spend (USD)"])
    for point in timeseries:
        trend_sheet.append(
            [point.date, point.calls, point.whatsapp_messages, point.qualified_count, point.usd_spend]
        )

    campaign_sheet = workbook.create_sheet("Campaign conversion")
    campaign_sheet.append(
        ["Name", "Channel", "Status", "Pending", "Calling", "Completed", "Failed", "Qualified", "Qualification rate"]
    )
    for row in campaign_rows:
        campaign_sheet.append(
            [
                row.name,
                row.channel,
                row.status,
                row.counts.pending,
                row.counts.calling,
                row.counts.completed,
                row.counts.failed,
                row.counts.qualified,
                row.qualification_rate if row.qualification_rate is not None else "",
            ]
        )

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    stamp = datetime.now(UTC).strftime("%Y-%m-%d")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="report-{stamp}.xlsx"'},
    )


@router.get("/conversations")
async def list_conversations(
    db: DbDep,
    scope_org_id: AnalyticsScopeDep,
    x_admin_token: str | None = Header(None),
    channel: str | None = Query(None, pattern="^(voice|whatsapp)$"),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
) -> list[dict]:
    msg_count_subq = (
        select(Message.conversation_id, func.count().label("message_count"))
        .group_by(Message.conversation_id)
        .subquery()
    )

    stmt = (
        select(Conversation, func.coalesce(msg_count_subq.c.message_count, 0), User.phone, User.name)
        .outerjoin(msg_count_subq, Conversation.id == msg_count_subq.c.conversation_id)
        .join(User, User.id == Conversation.user_id)
    )
    if scope_org_id is not None:
        stmt = stmt.where(Conversation.org_id == scope_org_id)
    if channel:
        stmt = stmt.where(Conversation.channel == channel)
    stmt = stmt.order_by(Conversation.started_at.desc()).limit(limit).offset(offset)

    rows = (await db.execute(stmt)).all()

    return [
        {
            **ConversationOut.model_validate(conv).model_dump(mode="json"),
            "message_count": int(count),
            "user_phone": phone,
            "user_name": name,
        }
        for conv, count, phone, name in rows
    ]


@router.get("/conversations/{conversation_id}/messages")
async def get_conversation_messages(
    conversation_id: UUID,
    db: DbDep,
    x_admin_token: str | None = Header(None),
) -> list[MessageOut]:
    stmt = (
        select(Message)
        .where(Message.conversation_id == conversation_id)
        .order_by(Message.created_at.asc())
    )
    messages = (await db.execute(stmt)).scalars().all()
    return [MessageOut.model_validate(m) for m in messages]


_SUMMARY_SYSTEM_PROMPT = (
    "Summarize this customer conversation in 2-3 sentences for a sales rep "
    "reviewing it later. Cover what the customer wanted, anything promised or "
    "decided, and any next step. Plain prose, no headers or bullet points."
)


@router.post("/conversations/{conversation_id}/summarize", response_model=ConversationOut)
async def summarize_conversation(
    conversation_id: UUID,
    db: DbDep,
    x_admin_token: str | None = Header(None),
) -> ConversationOut:
    """Generate (or regenerate) an AI summary of a conversation's transcript,
    on demand — not automatic, since not every conversation needs one."""
    conversation = await db.get(Conversation, conversation_id)
    if conversation is None:
        raise HTTPException(status_code=404, detail="Conversation not found")

    stmt = (
        select(Message)
        .where(Message.conversation_id == conversation_id)
        .order_by(Message.created_at.asc())
    )
    messages = (await db.execute(stmt)).scalars().all()
    transcript = "\n".join(f"{m.role}: {m.content}" for m in messages if m.content)
    if not transcript:
        raise HTTPException(status_code=400, detail="Conversation has no messages to summarize")

    result = await chat_completion(
        messages=[
            {"role": "system", "content": _SUMMARY_SYSTEM_PROMPT},
            {"role": "user", "content": transcript},
        ],
    )
    conversation.summary = (result.content or "").strip() or None
    await db.commit()
    await db.refresh(conversation)

    user = await db.get(User, conversation.user_id)
    return ConversationOut(
        **ConversationOut.model_validate(conversation).model_dump(exclude={"user_phone", "user_name"}),
        user_phone=user.phone if user else None,
        user_name=user.name if user else None,
    )


_LEAD_STATUS_PATTERN = f"^({'|'.join(LEAD_STATUSES)})$"
_LEAD_QUALIFICATION_STATUS_PATTERN = f"^({'|'.join(LEAD_QUALIFICATION_STATUSES)})$"


def _lead_tag_clause(tag: str):
    """Match a tag inside `Lead.tags` (a JSON string array) by casting to
    text and checking for the quoted value — portable across SQLite (tests)
    and Postgres (prod) without needing JSONB containment operators.
    """
    return cast(Lead.tags, String).ilike(f'%"{tag}"%')


def _lead_search_clause(search: str):
    """Unified search box (UI) — matches leads whose intent OR tags contain
    the term, so one field can stand in for the separate intent/tag filters.
    """
    return or_(Lead.intent.ilike(f"%{search}%"), cast(Lead.tags, String).ilike(f"%{search}%"))


def _lead_status_clause(status: str):
    """`status=qualified` also matches leads qualified via the separate
    qualification_status review workflow (db/models/lead.py) — a lead a rep
    marked qualified there commonly still sits at pipeline status
    "new"/"contacted". The leads-list filter (leads-view.tsx) offers a
    single "Qualified" option covering both rather than two near-duplicate
    ones, so this is the one status value where the filter means "either
    field says qualified" instead of an exact pipeline-status match.
    """
    if status == "qualified":
        return or_(Lead.status == "qualified", Lead.qualification_status == "qualified")
    return Lead.status == status


@router.get("/leads")
async def list_leads(
    db: DbDep,
    scope_org_id: AnalyticsScopeDep,
    x_admin_token: str | None = Header(None),
    intent: str | None = Query(None),
    channel: str | None = Query(None, pattern="^(voice|whatsapp)$"),
    status: str | None = Query(None, pattern=_LEAD_STATUS_PATTERN),
    qualification_status: str | None = Query(None, pattern=_LEAD_QUALIFICATION_STATUS_PATTERN),
    tag: str | None = Query(None, description="Filter by a single tag"),
    search: str | None = Query(None, description="Match against intent or tags"),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
) -> list[LeadOut]:
    stmt = select(Lead).order_by(Lead.created_at.desc()).limit(limit).offset(offset)
    if scope_org_id is not None:
        stmt = stmt.where(Lead.org_id == scope_org_id)
    if intent:
        stmt = stmt.where(Lead.intent.ilike(f"%{intent}%"))
    if channel:
        stmt = stmt.where(Lead.channel == channel)
    if status:
        stmt = stmt.where(_lead_status_clause(status))
    if qualification_status:
        stmt = stmt.where(Lead.qualification_status == qualification_status)
    if tag:
        stmt = stmt.where(_lead_tag_clause(tag))
    if search:
        stmt = stmt.where(_lead_search_clause(search))

    leads = (await db.execute(stmt)).scalars().all()
    return [LeadOut.model_validate(lead) for lead in leads]


_SAMPLE_IMPORT_ROWS = [
    {"name": "Asha Verma", "phone": "+919876543210", "call": "yes", "whatsapp": "no", "status": "new"},
    {"name": "Rohit Singh", "phone": "+919812345678", "call": "no", "whatsapp": "yes", "status": "contacted"},
    {"name": "Priya Nair", "phone": "+919845098450", "call": "yes", "whatsapp": "yes", "status": "qualified"},
]


def _csv_streaming_response(csv_text: str, filename: str) -> StreamingResponse:
    """Wrap CSV text as a downloadable response with a UTF-8 BOM prefix.

    Without the BOM, Excel on Windows guesses the file's encoding from the
    system codepage instead of UTF-8 — any non-ASCII character (accented
    names, curly quotes, ...) renders as mojibake even though the bytes
    themselves are valid UTF-8. The BOM makes Excel detect UTF-8 correctly;
    other tools (Python's csv module, Google Sheets, ...) already ignore or
    strip a leading BOM, so this doesn't break anything else that reads it.
    """
    return StreamingResponse(
        iter(["﻿" + csv_text]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/leads/sample.csv")
async def sample_leads_csv(x_admin_token: str | None = Header(None)) -> StreamingResponse:
    """Blank-data template for POST /leads/import — same columns that
    endpoint reads (name, phone, call, whatsapp, status), so a unified-page
    upload can mix call-only, WhatsApp-only, and both-channel rows in one
    file. `status` is required here (unlike POST /admin/campaigns' sample,
    which has no such column) since this endpoint creates a Lead per row
    immediately — see import_leads_file's required_columns.
    Registered ahead of GET /leads/{lead_id} so its literal path isn't
    swallowed by that route's UUID path param.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["name", "phone", "call", "whatsapp", "status"])
    for row in _SAMPLE_IMPORT_ROWS:
        writer.writerow([row["name"], row["phone"], row["call"], row["whatsapp"], row["status"]])
    buf.seek(0)

    return _csv_streaming_response(buf.getvalue(), "leads-sample.csv")


@router.get("/leads/sample.xlsx")
async def sample_leads_xlsx(x_admin_token: str | None = Header(None)) -> StreamingResponse:
    """Same template as GET /leads/sample.csv, as an .xlsx workbook."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    sheet.append(["name", "phone", "call", "whatsapp", "status"])
    for row in _SAMPLE_IMPORT_ROWS:
        sheet.append([row["name"], row["phone"], row["call"], row["whatsapp"], row["status"]])
    for cell in sheet["B"][1:]:
        cell.number_format = "@"

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="leads-sample.xlsx"'},
    )


@router.get("/leads/{lead_id}", response_model=LeadDetailOut)
async def get_lead(
    lead_id: UUID,
    db: DbDep,
    x_admin_token: str | None = Header(None),
) -> LeadDetailOut:
    """Lead detail — the captured fields plus that lead's conversation
    history (dashboard/CRM view). Conversations are joined via the shared
    ``user_id`` since Lead has no direct FK to Conversation.
    """
    lead = (await db.execute(select(Lead).where(Lead.id == lead_id))).scalar_one_or_none()
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")

    msg_count_subq = (
        select(Message.conversation_id, func.count().label("message_count"))
        .group_by(Message.conversation_id)
        .subquery()
    )
    conv_stmt = (
        select(Conversation, func.coalesce(msg_count_subq.c.message_count, 0), User.phone, User.name)
        .outerjoin(msg_count_subq, Conversation.id == msg_count_subq.c.conversation_id)
        .join(User, User.id == Conversation.user_id)
        .where(Conversation.user_id == lead.user_id)
        .order_by(Conversation.started_at.desc())
    )
    conv_rows = (await db.execute(conv_stmt)).all()
    conversations = [
        ConversationSummaryOut(
            **ConversationOut.model_validate(conv).model_dump(exclude={"user_phone", "user_name"}),
            message_count=int(count),
            user_phone=phone,
            user_name=name,
        )
        for conv, count, phone, name in conv_rows
    ]

    return LeadDetailOut(**LeadOut.model_validate(lead).model_dump(), conversations=conversations)


@router.patch("/leads/{lead_id}", response_model=LeadOut)
async def update_lead(
    lead_id: UUID,
    payload: LeadUpdateIn,
    db: DbDep,
    x_admin_token: str | None = Header(None),
) -> LeadOut:
    """Update a lead's status and/or follow-up. Partial: only fields the
    caller actually sent are touched, so posting `{"status": "contacted"}`
    leaves follow_up_at/note untouched, while `{"follow_up_at": null}`
    explicitly clears it.
    """
    lead = (await db.execute(select(Lead).where(Lead.id == lead_id))).scalar_one_or_none()
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")

    updates = payload.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(lead, field, value)

    # Stamp qualified_at the moment a lead first enters "qualified", so the
    # caller doesn't have to set it explicitly on every qualification update.
    if updates.get("qualification_status") == "qualified" and lead.qualified_at is None:
        lead.qualified_at = datetime.now(UTC)

    await db.commit()
    await db.refresh(lead)
    return LeadOut.model_validate(lead)


def _leads_export_stmt(
    scope_org_id: UUID | None,
    intent: str | None,
    channel: str | None,
    status: str | None,
    qualification_status: str | None,
    tag: str | None,
    search: str | None,
    limit: int,
    offset: int,
):
    stmt = select(Lead).order_by(Lead.created_at.desc()).limit(limit).offset(offset)
    if scope_org_id is not None:
        stmt = stmt.where(Lead.org_id == scope_org_id)
    if intent:
        stmt = stmt.where(Lead.intent.ilike(f"%{intent}%"))
    if channel:
        stmt = stmt.where(Lead.channel == channel)
    if status:
        stmt = stmt.where(_lead_status_clause(status))
    if qualification_status:
        stmt = stmt.where(Lead.qualification_status == qualification_status)
    if tag:
        stmt = stmt.where(_lead_tag_clause(tag))
    if search:
        stmt = stmt.where(_lead_search_clause(search))
    return stmt


_LEADS_EXPORT_HEADER = [
    "id",
    "name",
    "phone",
    "intent",
    "tags",
    "channel",
    "status",
    "qualification_status",
    "qualification_score",
    "created_at",
]


def _lead_export_row(lead: Lead) -> list[str]:
    return [
        str(lead.id),
        lead.name or "",
        lead.phone or "",
        lead.intent or "",
        ",".join(lead.tags) if lead.tags else "",
        lead.channel or "",
        lead.status or "",
        lead.qualification_status or "",
        str(lead.qualification_score) if lead.qualification_score is not None else "",
        lead.created_at.isoformat() if lead.created_at else "",
    ]


@router.get("/leads.csv")
async def export_leads_csv(
    db: DbDep,
    scope_org_id: AnalyticsScopeDep,
    x_admin_token: str | None = Header(None),
    intent: str | None = Query(None),
    channel: str | None = Query(None, pattern="^(voice|whatsapp)$"),
    status: str | None = Query(None, pattern=_LEAD_STATUS_PATTERN),
    qualification_status: str | None = Query(None, pattern=_LEAD_QUALIFICATION_STATUS_PATTERN),
    tag: str | None = Query(None, description="Filter by a single tag"),
    search: str | None = Query(None, description="Match against intent or tags"),
    limit: int = Query(1000, ge=1, le=10000),
    offset: int = Query(0, ge=0),
) -> StreamingResponse:
    """Same data as GET /admin/leads but rendered as CSV for download —
    including the same org scoping, so an export can't pull rows the list
    view wouldn't show."""
    stmt = _leads_export_stmt(
        scope_org_id, intent, channel, status, qualification_status, tag, search, limit, offset
    )
    leads = (await db.execute(stmt)).scalars().all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_LEADS_EXPORT_HEADER)
    for lead in leads:
        writer.writerow(_lead_export_row(lead))
    buf.seek(0)

    return _csv_streaming_response(buf.getvalue(), "leads.csv")


@router.get("/leads.xlsx")
async def export_leads_xlsx(
    db: DbDep,
    scope_org_id: AnalyticsScopeDep,
    x_admin_token: str | None = Header(None),
    intent: str | None = Query(None),
    channel: str | None = Query(None, pattern="^(voice|whatsapp)$"),
    status: str | None = Query(None, pattern=_LEAD_STATUS_PATTERN),
    qualification_status: str | None = Query(None, pattern=_LEAD_QUALIFICATION_STATUS_PATTERN),
    tag: str | None = Query(None, description="Filter by a single tag"),
    search: str | None = Query(None, description="Match against intent or tags"),
    limit: int = Query(1000, ge=1, le=10000),
    offset: int = Query(0, ge=0),
) -> StreamingResponse:
    """Same data/filters as GET /admin/leads.csv, as an .xlsx workbook —
    e.g. `?qualification_status=qualified` for the Reports page's "Export
    qualified leads" button."""
    stmt = _leads_export_stmt(
        scope_org_id, intent, channel, status, qualification_status, tag, search, limit, offset
    )
    leads = (await db.execute(stmt)).scalars().all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    sheet.append(_LEADS_EXPORT_HEADER)
    for lead in leads:
        sheet.append(_lead_export_row(lead))
    for cell in sheet["C"][1:]:  # phone column — keep leading "+"/zeros as text
        cell.number_format = "@"

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    filename = "qualified-leads.xlsx" if qualification_status == "qualified" else "leads.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _normalize_import_row(raw_row: dict[str, object]) -> dict[str, str]:
    return {
        (k or "").strip().lower(): (str(v).strip() if v is not None else "")
        for k, v in raw_row.items()
    }


def _parse_bool_cell(value: str) -> bool:
    return value.strip().lower() in {"1", "y", "yes", "true"}


def _resolve_row_channels(row: dict[str, str]) -> list[str] | None:
    """Per-row channel resolution for unified (multi-channel) uploads.

    Priority: the new 'call'/'whatsapp' boolean-ish columns (a row may
    resolve to one, both, or neither); else the legacy single 'channel' text
    column (voice XOR whatsapp), kept for backward compatibility.

    Returns ``None`` when the row gives no usable signal at all — the caller
    must treat this as a visible per-row error, NOT default to voice (that
    silent default was the root cause of "WhatsApp list uploaded but nothing
    sent").
    """
    if "call" in row or "whatsapp" in row:
        channels = []
        if row.get("call") and _parse_bool_cell(row["call"]):
            channels.append("voice")
        if row.get("whatsapp") and _parse_bool_cell(row["whatsapp"]):
            channels.append("whatsapp")
        return channels or None
    if "channel" in row:
        value = row["channel"].strip().lower()
        return [value] if value in ("voice", "whatsapp") else None
    return None


def _parse_template_params_form(raw: str | None) -> list[str] | None:
    """Decode the JSON-encoded ``template_params`` Form field (a list of
    per-placeholder tokens/literals — see CallCampaign.template_params) sent
    by the Campaigns-page upload form. ``None``/blank means no template
    params configured, distinct from an (invalid) empty JSON array."""
    error_detail = "template_params must be a JSON array of strings"
    if not raw:
        return None
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, TypeError) as exc:
        raise HTTPException(status_code=400, detail=error_detail) from exc
    if not isinstance(parsed, list) or not all(isinstance(v, str) for v in parsed):
        raise HTTPException(status_code=400, detail=error_detail)
    return parsed


def _decode_csv_bytes(raw: bytes) -> str:
    """Best-effort decode of an uploaded CSV's raw bytes.

    UTF-8 (with or without a BOM) is tried first since it's the correct,
    unambiguous case. Falls back to cp1252 — what Excel on Windows saves as
    by default, and the single most common reason a real user's CSV isn't
    valid UTF-8 (curly quotes, accented names, the Euro sign, ...) — before
    giving up with a clear error instead of the raw UnicodeDecodeError
    previously leaking out of this as an unhandled 500.
    """
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            return raw.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Could not read this CSV's text encoding. Please save/export it as "
                    "UTF-8 CSV and try again."
                ),
            ) from exc


def _missing_columns_detail(missing: set[str]) -> str:
    cols = sorted(missing)
    if len(cols) == 1:
        return f"File must include a '{cols[0]}' column"
    quoted = ", ".join(f"'{c}'" for c in cols)
    return f"File must include the following columns: {quoted}"


def _iter_csv_rows(
    raw: bytes, required_columns: frozenset[str] = frozenset({"phone"})
) -> Iterator[tuple[int, dict[str, str]]]:
    reader = csv.DictReader(io.StringIO(_decode_csv_bytes(raw)))
    headers = {(f or "").strip().lower() for f in (reader.fieldnames or [])}
    missing = required_columns - headers
    if missing:
        raise HTTPException(status_code=400, detail=_missing_columns_detail(missing))
    for row_num, raw_row in enumerate(reader, start=2):  # header occupies row 1
        yield row_num, _normalize_import_row(raw_row)


def _iter_xlsx_rows(
    raw: bytes, required_columns: frozenset[str] = frozenset({"phone"})
) -> Iterator[tuple[int, dict[str, str]]]:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Could not read .xlsx file") from exc

    sheet_rows = workbook.active.iter_rows(values_only=True)
    header = next(sheet_rows, None) or ()
    headers = [(str(h).strip().lower() if h is not None else "") for h in header]
    missing = required_columns - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=_missing_columns_detail(missing))

    for row_num, values in enumerate(sheet_rows, start=2):  # header occupies row 1
        yield row_num, _normalize_import_row(dict(zip(headers, values, strict=False)))


def _default_campaign_name() -> str:
    return f"Bulk import {datetime.now(UTC):%Y-%m-%d %H:%M}"


# Default qualification bar for campaigns auto-created from a plain lead
# import, when the caller doesn't supply their own ``criteria``.
_DEFAULT_IMPORT_CRITERIA = (
    "Prospect confirms genuine interest in our product/service and provides "
    "valid contact details for follow-up."
)


@router.post("/leads/import", response_model=CampaignCreateResult)
async def import_leads_file(
    db: DbDep,
    org: RequestOrgDep,
    file: UploadFile = File(...),
    x_admin_token: str | None = Header(None),
    channel: str | None = Query(None, pattern="^(voice|whatsapp)$"),
    campaign_name: str | None = Form(None),
    criteria: str | None = Form(None),
    start_mode: str = Form("draft", pattern="^(draft|now|scheduled)$"),
    scheduled_start_at: datetime | None = Form(None),
    template_name: str | None = Form(None),
    template_language: str | None = Form(None),
    template_params: str | None = Form(None),
    custom_message: str | None = Form(None),
) -> CampaignCreateResult:
    """Bulk-import leads from an uploaded CSV or Excel (.xlsx) file.

    Imported contacts are staged as an auto-generated campaign
    (``CallCampaign`` + ``CampaignTarget`` rows), same as ``POST
    /admin/campaigns`` — the background dialer/dispatcher still reaches out
    to each one. Unlike that endpoint, though, this one ALSO writes a
    ``Lead`` row immediately for every contact (``auto_qualify=True``, see
    ``_create_campaign_from_rows``) rather than waiting for the AI's
    ``qualify_lead`` tool call, and it fills in a default campaign
    name/criteria when the caller doesn't supply one. For programmatic
    (non-file) bulk import, see ``POST /admin/leads/bulk``.

    ``channel`` is optional: per-channel pages (e.g. /calling/leads,
    /whatsapp/leads) pass it explicitly, forcing every row into that one
    channel. The unified CRM leads page omits it, since a single upload may
    mix call and WhatsApp rows — in that case each row's own ``call``/
    ``whatsapp`` columns (or legacy ``channel`` column) pick its
    destination(s) and the file is split into one campaign per channel found.

    The file must also include a ``status`` column (one of "new",
    "contacted", "qualified", "converted", "lost") — required, with no
    default, since it becomes each row's immediately-created Lead's status.
    A blank or invalid cell is a per-row import error, same as a bad phone
    number. This is the pipeline status only — it doesn't touch
    ``qualification_status`` (the separate rep-review workflow), which stays
    at its default and is edited per-lead afterward from the Leads page.

    ``start_mode`` controls whether the resulting campaign(s) begin outreach
    immediately ("now"), sit inert until manually started ("draft", the
    default — uploading a list never starts calling/messaging on its own),
    or begin at a future ``scheduled_start_at`` ("scheduled").
    """
    if start_mode == "scheduled" and (
        scheduled_start_at is None or scheduled_start_at <= datetime.now(UTC)
    ):
        raise HTTPException(status_code=400, detail="scheduled_start_at must be in the future")
    parsed_template_params = _parse_template_params_form(template_params)

    filename = (file.filename or "").lower()
    raw = await file.read()

    # A Lead is created immediately for every row here (auto_qualify=True
    # below) with that row's own status, so — unlike POST /admin/campaigns,
    # which never touches Lead.status — the file must supply one per row
    # rather than defaulting silently to "qualified".
    required_columns = frozenset({"phone", "status"})
    if filename.endswith(".csv"):
        rows = list(_iter_csv_rows(raw, required_columns))
    elif filename.endswith(".xlsx"):
        rows = list(_iter_xlsx_rows(raw, required_columns))
    else:
        raise HTTPException(status_code=400, detail="Only .csv or .xlsx files are supported")

    org_id = org
    name = campaign_name or _default_campaign_name()
    crit = criteria or _DEFAULT_IMPORT_CRITERIA
    resolved_status = {"draft": "draft", "now": "running", "scheduled": "scheduled"}[start_mode]

    return await _create_campaigns_from_rows(
        db,
        org_id=org_id,
        name=name,
        criteria=crit,
        forced_channel=channel,
        rows=rows,
        auto_qualify=True,
        status=resolved_status,
        scheduled_start_at=scheduled_start_at if start_mode == "scheduled" else None,
        template_name=template_name,
        template_language=template_language,
        template_params=parsed_template_params,
        custom_message=custom_message,
    )


@router.post("/leads/bulk", response_model=CampaignCreateResult)
async def import_leads_bulk(
    payload: LeadBulkImportIn,
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> CampaignCreateResult:
    """Bulk-import leads from a JSON array — the programmatic counterpart to
    the CSV/Excel upload above, for callers integrating directly against the
    API rather than uploading a file. Same auto-campaign behavior as
    ``POST /admin/leads/import``.

    ``channel``, when set, forces every row into that one channel (unchanged
    behavior for existing integrations). When omitted, rows may instead set
    their own ``call``/``whatsapp`` flags for per-row routing — but only if
    at least one row actually uses them; if none do, this still defaults to
    "voice" exactly as before, so existing callers see no behavior change.
    """
    if payload.start_mode == "scheduled" and (
        payload.scheduled_start_at is None or payload.scheduled_start_at <= datetime.now(UTC)
    ):
        raise HTTPException(status_code=400, detail="scheduled_start_at must be in the future")

    org_id = org
    any_row_channel_set = any(
        row.call is not None or row.whatsapp is not None for row in payload.leads
    )
    if payload.channel:
        forced_channel: str | None = payload.channel
    elif any_row_channel_set:
        forced_channel = None
    else:
        forced_channel = "voice"

    def _row_dict(row: LeadImportRow) -> dict[str, str]:
        d = {"phone": row.phone, "name": row.name or ""}
        if row.call is not None:
            d["call"] = "yes" if row.call else "no"
        if row.whatsapp is not None:
            d["whatsapp"] = "yes" if row.whatsapp else "no"
        return d

    rows = [(row_num, _row_dict(row)) for row_num, row in enumerate(payload.leads, start=1)]
    resolved_status = {"draft": "draft", "now": "running", "scheduled": "scheduled"}[
        payload.start_mode
    ]

    return await _create_campaigns_from_rows(
        db,
        org_id=org_id,
        name=payload.campaign_name or _default_campaign_name(),
        criteria=payload.criteria or _DEFAULT_IMPORT_CRITERIA,
        forced_channel=forced_channel,
        auto_qualify=True,
        rows=rows,
        status=resolved_status,
        scheduled_start_at=payload.scheduled_start_at if payload.start_mode == "scheduled" else None,
        template_name=payload.template_name,
        template_language=payload.template_language,
        template_params=payload.template_params,
        custom_message=payload.custom_message,
    )


# ---------------------------------------------------------------------------
# Campaigns — bulk-upload a lead list, auto-reach each one by voice or
# WhatsApp, and let the AI qualify them (core/tools.qualify_lead) against
# per-campaign criteria. Only qualified targets ever produce a Lead row.
# Every bulk-import entry point (this section's /campaigns, plus
# /leads/import and /leads/bulk above) funnels through
# _create_campaign_from_rows so none of them can bypass qualification.
# ---------------------------------------------------------------------------

# Plivo dials whatever's in `to` verbatim — a bare 10-digit number without a
# country code won't ring (mirrors the Dial page's zod regex, apps/web's
# calling/dial/page.tsx), so campaign uploads are validated the same way.
_E164_PATTERN = re.compile(r"^\+\d{8,15}$")


async def _campaign_counts(db: DbDep, campaign_id: UUID) -> CampaignCounts:
    counts = await _campaign_counts_bulk(db, [campaign_id])
    return counts.get(campaign_id, CampaignCounts(pending=0, calling=0, completed=0, failed=0, qualified=0))


async def _campaign_counts_bulk(
    db: DbDep, campaign_ids: Sequence[UUID]
) -> dict[UUID, CampaignCounts]:
    """Batched counts for multiple campaigns in 1 round trip total instead of
    one query per campaign (or even 2 for all of them) — used by list/report
    endpoints that render a row per campaign. Status tallies and the
    qualified count both come off the same ``campaign_targets`` rows, so a
    single grouped query with conditional SUMs gets both instead of two
    separate round trips to the same table (DB is a remote Neon instance —
    see ``get_stats`` above for the same pattern)."""
    if not campaign_ids:
        return {}

    def _count_if(condition):  # noqa: ANN001, ANN202 — SQLAlchemy boolean expr
        return func.coalesce(func.sum(case((condition, 1), else_=0)), 0)

    combined_stmt = (
        select(
            CampaignTarget.campaign_id,
            _count_if(CampaignTarget.status == "pending"),
            _count_if(CampaignTarget.status == "calling"),
            _count_if(CampaignTarget.status == "completed"),
            _count_if(CampaignTarget.status == "failed"),
            _count_if(CampaignTarget.qualified.is_(True)),
        )
        .where(CampaignTarget.campaign_id.in_(campaign_ids))
        .group_by(CampaignTarget.campaign_id)
    )
    counts_by_id = {
        cid: CampaignCounts(pending=0, calling=0, completed=0, failed=0, qualified=0)
        for cid in campaign_ids
    }
    for cid, pending, calling, completed, failed, qualified in (
        await db.execute(combined_stmt)
    ).all():
        counts_by_id[cid] = CampaignCounts(
            pending=pending,
            calling=calling,
            completed=completed,
            failed=failed,
            qualified=qualified,
        )
    return counts_by_id


def _campaign_out(campaign: CallCampaign, counts: CampaignCounts) -> CampaignOut:
    return CampaignOut(
        id=campaign.id,
        org_id=campaign.org_id,
        name=campaign.name,
        criteria=campaign.criteria,
        channel=campaign.channel,
        status=campaign.status,
        scheduled_start_at=campaign.scheduled_start_at,
        template_name=campaign.template_name,
        template_language=campaign.template_language,
        template_params=campaign.template_params,
        custom_message=campaign.custom_message,
        created_at=campaign.created_at,
        counts=counts,
    )


async def _create_campaign_from_rows(
    db: DbDep,
    *,
    org_id: UUID,
    name: str,
    criteria: str,
    rows: Iterable[tuple[int, dict[str, str], str]],
    auto_qualify: bool = False,
    status: str = "draft",
    scheduled_start_at: datetime | None = None,
    template_name: str | None = None,
    template_language: str | None = None,
    template_params: list[str] | None = None,
    custom_message: str | None = None,
) -> CampaignCreateResult:
    """Shared core for every campaign-creation entry point (CSV/xlsx upload,
    JSON bulk import). Creates exactly ONE ``CallCampaign`` regardless of how
    many distinct channels appear across ``rows`` — a row already resolved to
    both call and whatsapp appears twice in ``rows`` (once per channel) and
    produces two ``CampaignTarget`` rows under this single campaign, one per
    channel. ``CallCampaign.channel`` is set once at the end from the
    distinct channels actually inserted ("voice", "whatsapp", or "mixed") —
    it's a display summary only; routing is per-target (CampaignTarget.channel).

    By default (``auto_qualify=False``, the Campaigns page's ``POST
    /admin/campaigns``) contacts are staged as ``CampaignTarget`` rows only —
    the background dialer/dispatcher calls or messages each one and the AI's
    ``qualify_lead`` tool call is what promotes a target into the CRM leads
    table, and only when interested. Campaign behavior itself is unchanged
    either way.

    ``auto_qualify=True`` (the leads-page entry points, ``POST
    /admin/leads/import`` and ``POST /admin/leads/bulk``) additionally
    creates a ``Lead`` row immediately for every successfully staged
    contact — an org uploading a list of contacts they already trust
    shouldn't have to wait for the AI to call each one before seeing them on
    the Leads page. The org can change that status afterward like any other
    lead. A row resolved to both channels still produces one Lead per
    channel, matching its two outreach attempts.

    That Lead's status comes from the row's own ``status`` column when one
    was supplied (required for the CSV/xlsx upload — see
    ``import_leads_file``'s ``required_columns``; an invalid value is a
    per-row error, same as a bad phone number). ``POST /admin/leads/bulk``'s
    JSON rows have no such column and keep the old default of "qualified".
    """
    campaign = CallCampaign(
        org_id=org_id,
        name=name,
        criteria=criteria,
        channel="voice",  # placeholder — corrected below once targets are known
        status=status,
        scheduled_start_at=scheduled_start_at,
        template_name=template_name,
        template_language=template_language,
        template_params=template_params,
        custom_message=custom_message,
        created_at=datetime.now(UTC),
    )
    db.add(campaign)
    await db.flush()

    imported = 0
    errors: list[dict[str, str | int]] = []
    seen_channels: set[str] = set()
    for row_num, row, channel in rows:
        phone = row.get("phone", "")
        if not phone:
            errors.append({"row": row_num, "reason": "missing phone"})
            continue
        normalized = _normalize_phone(phone)
        if not _E164_PATTERN.match(normalized):
            errors.append(
                {
                    "row": row_num,
                    "reason": (
                        f"phone '{phone}' must include a country code in E.164 format, "
                        "e.g. +919876543210"
                    ),
                }
            )
            continue
        row_name = row.get("name") or None
        # "status" is only present at all for file-based rows (import_leads_file
        # requires the column via _iter_csv_rows/_iter_xlsx_rows's
        # required_columns); JSON bulk-import rows (_row_dict) never set the
        # key, so they keep the old default of "qualified" untouched. This is
        # the pipeline status only — qualification_status (the separate
        # rep-review workflow) isn't settable from the import file; it stays
        # at its default ("unqualified") and is edited per-lead afterward.
        row_status = "qualified"
        if auto_qualify and "status" in row:
            candidate = row.get("status", "").strip().lower()
            if candidate not in LEAD_STATUSES:
                errors.append(
                    {
                        "row": row_num,
                        "reason": f"status must be one of: {', '.join(LEAD_STATUSES)}",
                    }
                )
                continue
            row_status = candidate
        db.add(
            CampaignTarget(
                campaign_id=campaign.id,
                org_id=org_id,
                name=row_name,
                phone=normalized,
                channel=channel,
            )
        )
        if auto_qualify:
            user = await _get_or_create_user_by_phone(db, org_id, normalized, row_name)
            db.add(
                Lead(
                    org_id=org_id,
                    user_id=user.id,
                    name=row_name,
                    phone=normalized,
                    intent="imported",
                    channel=channel,
                    status=row_status,
                )
            )
        seen_channels.add(channel)
        imported += 1

    campaign.channel = seen_channels.pop() if len(seen_channels) == 1 else (
        "mixed" if len(seen_channels) > 1 else "voice"
    )

    await db.commit()

    logger.info(
        "admin_campaign_created",
        campaign_id=str(campaign.id),
        channel=campaign.channel,
        imported=imported,
        skipped=len(errors),
    )

    counts = await _campaign_counts(db, campaign.id)
    campaign_out = _campaign_out(campaign, counts)
    return CampaignCreateResult(
        campaign=campaign_out,
        campaigns=[campaign_out],
        imported=imported,
        skipped=len(errors),
        errors=errors,
    )


async def _create_campaigns_from_rows(
    db: DbDep,
    *,
    org_id: UUID,
    name: str,
    criteria: str,
    forced_channel: str | None,
    rows: list[tuple[int, dict[str, str]]],
    auto_qualify: bool = False,
    status: str = "draft",
    scheduled_start_at: datetime | None = None,
    template_name: str | None = None,
    template_language: str | None = None,
    template_params: list[str] | None = None,
    custom_message: str | None = None,
) -> CampaignCreateResult:
    """Multi-channel orchestrator on top of ``_create_campaign_from_rows``.
    Always creates exactly ONE campaign, regardless of how many distinct
    channels the upload's rows resolve to.

    When ``forced_channel`` is given (an explicit per-channel page, or a
    caller that already knows its channel), every row goes to that one
    channel unchanged — row-level call/whatsapp/channel columns are ignored.

    When ``forced_channel`` is ``None``, each row is resolved independently
    via ``_resolve_row_channels`` and may resolve to BOTH voice and
    whatsapp — that row is expanded into two ``(row_num, row, channel)``
    entries, producing two ``CampaignTarget`` rows under the same campaign.
    Rows that resolve to neither channel are recorded as a visible per-row
    error instead of silently defaulting to voice.
    """
    expanded: list[tuple[int, dict[str, str], str]] = []
    row_errors: list[dict[str, str | int]] = []
    for row_num, row in rows:
        if forced_channel:
            expanded.append((row_num, row, forced_channel))
            continue
        channels = _resolve_row_channels(row)
        if not channels:
            row_errors.append(
                {
                    "row": row_num,
                    "reason": (
                        "row must mark 'call' and/or 'whatsapp' as yes, or set a "
                        "'channel' column to 'voice'/'whatsapp' — this row was not "
                        "sent on any channel"
                    ),
                }
            )
            continue
        for ch in channels:
            expanded.append((row_num, row, ch))

    if not expanded:
        raise HTTPException(
            status_code=400,
            detail=(
                "No row could be routed to a channel — add 'call'/'whatsapp' "
                "columns (yes/no) or a 'channel' column (voice/whatsapp)."
            ),
        )

    result = await _create_campaign_from_rows(
        db,
        org_id=org_id,
        name=name,
        criteria=criteria,
        rows=expanded,
        auto_qualify=auto_qualify,
        status=status,
        scheduled_start_at=scheduled_start_at,
        template_name=template_name,
        template_language=template_language,
        template_params=template_params,
        custom_message=custom_message,
    )
    result.errors = row_errors + result.errors
    result.skipped += len(row_errors)
    return result


@router.post("/campaigns", response_model=CampaignCreateResult)
async def create_campaign(
    db: DbDep,
    org: RequestOrgDep,
    name: str = Form(...),
    criteria: str = Form(...),
    file: UploadFile = File(...),
    channel: str | None = Form(None, pattern="^(voice|whatsapp)$"),
    start_mode: str = Form("draft", pattern="^(draft|now|scheduled)$"),
    scheduled_start_at: datetime | None = Form(None),
    template_name: str | None = Form(None),
    template_language: str | None = Form(None),
    template_params: str | None = Form(None),
    custom_message: str | None = Form(None),
    x_admin_token: str | None = Header(None),
) -> CampaignCreateResult:
    """Create a campaign from an uploaded CSV/Excel contact list.

    ``channel``, when given, forces every row into that one worker: the
    voice dialer (apps/api/workers/campaign_dialer.py) or the WhatsApp
    dispatcher (apps/api/workers/whatsapp_dispatcher.py). When omitted (the
    web app's Campaigns page never sends it), each row's own ``call``/
    ``whatsapp`` columns decide its channel(s) instead — see
    ``_create_campaigns_from_rows``.

    ``start_mode`` controls whether the resulting campaign(s) begin outreach
    immediately, sit as a draft until manually started (the default), or
    begin at a future ``scheduled_start_at``.
    """
    if start_mode == "scheduled" and (
        scheduled_start_at is None or scheduled_start_at <= datetime.now(UTC)
    ):
        raise HTTPException(status_code=400, detail="scheduled_start_at must be in the future")
    parsed_template_params = _parse_template_params_form(template_params)

    filename = (file.filename or "").lower()
    raw = await file.read()
    if filename.endswith(".csv"):
        rows = list(_iter_csv_rows(raw))
    elif filename.endswith(".xlsx"):
        rows = list(_iter_xlsx_rows(raw))
    else:
        raise HTTPException(status_code=400, detail="Only .csv or .xlsx files are supported")

    org_id = org
    usage = await get_credit_usage(db, org_id)
    await enforce_plan_limit(db, org_id, "max_campaigns", usage.campaigns)
    resolved_status = {"draft": "draft", "now": "running", "scheduled": "scheduled"}[start_mode]
    return await _create_campaigns_from_rows(
        db,
        org_id=org_id,
        name=name,
        criteria=criteria,
        forced_channel=channel,
        rows=rows,
        status=resolved_status,
        scheduled_start_at=scheduled_start_at if start_mode == "scheduled" else None,
        template_name=template_name,
        template_language=template_language,
        template_params=parsed_template_params,
        custom_message=custom_message,
    )


@router.get("/campaigns", response_model=list[CampaignOut])
async def list_campaigns(
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
    channel: str | None = Query(None, pattern="^(voice|whatsapp)$"),
) -> list[CampaignOut]:
    """List campaigns with their target-status counts in a single round
    trip — the counts come along as correlated subqueries on the same
    SELECT instead of a separate query afterward (DB is a remote Neon
    instance; see ``get_stats`` above for the same pattern applied to the
    dashboard stats tile)."""
    org_id = org

    def _target_count(condition):  # noqa: ANN001, ANN202 — SQLAlchemy boolean expr
        return (
            select(func.coalesce(func.sum(case((condition, 1), else_=0)), 0))
            .where(CampaignTarget.campaign_id == CallCampaign.id)
            .correlate(CallCampaign)
            .scalar_subquery()
        )

    stmt = (
        select(
            CallCampaign,
            _target_count(CampaignTarget.status == "pending"),
            _target_count(CampaignTarget.status == "calling"),
            _target_count(CampaignTarget.status == "completed"),
            _target_count(CampaignTarget.status == "failed"),
            _target_count(CampaignTarget.qualified.is_(True)),
        )
        .where(CallCampaign.org_id == org_id)
        .order_by(CallCampaign.created_at.desc())
    )
    if channel:
        # CallCampaign.channel is a display-only summary ("voice"/"whatsapp"/
        # "mixed") — filtering on it directly would wrongly exclude "mixed"
        # campaigns that do contain a target of the requested channel.
        stmt = stmt.where(
            select(CampaignTarget.id)
            .where(CampaignTarget.campaign_id == CallCampaign.id, CampaignTarget.channel == channel)
            .exists()
        )

    rows = (await db.execute(stmt)).all()
    return [
        _campaign_out(
            campaign,
            CampaignCounts(pending=pending, calling=calling, completed=completed, failed=failed, qualified=qualified),
        )
        for campaign, pending, calling, completed, failed, qualified in rows
    ]


_SAMPLE_CAMPAIGN_ROWS = _SAMPLE_IMPORT_ROWS


@router.get("/campaigns/sample.csv")
async def sample_campaign_csv(x_admin_token: str | None = Header(None)) -> StreamingResponse:
    """Blank-data template for POST /campaigns' contact-list upload — each
    row picks its own channel(s) via the call/whatsapp columns, so one list
    can mix call-only, WhatsApp-only, and both-channel contacts. Registered
    ahead of GET /campaigns/{campaign_id} so its literal path isn't
    swallowed by that route's UUID path param.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["name", "phone", "call", "whatsapp"])
    for row in _SAMPLE_CAMPAIGN_ROWS:
        writer.writerow([row["name"], row["phone"], row["call"], row["whatsapp"]])
    buf.seek(0)

    return _csv_streaming_response(buf.getvalue(), "campaign-contacts-sample.csv")


@router.get("/campaigns/sample.xlsx")
async def sample_campaign_xlsx(x_admin_token: str | None = Header(None)) -> StreamingResponse:
    """Same template as GET /campaigns/sample.csv, as an .xlsx workbook."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"
    sheet.append(["name", "phone", "call", "whatsapp"])
    for row in _SAMPLE_CAMPAIGN_ROWS:
        sheet.append([row["name"], row["phone"], row["call"], row["whatsapp"]])
    for cell in sheet["B"][1:]:
        cell.number_format = "@"

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="campaign-contacts-sample.xlsx"'},
    )


@router.get("/campaigns/{campaign_id}", response_model=CampaignDetailOut)
async def get_campaign(
    campaign_id: UUID,
    db: DbDep,
    x_admin_token: str | None = Header(None),
) -> CampaignDetailOut:
    campaign = await db.get(CallCampaign, campaign_id)
    if campaign is None:
        raise HTTPException(status_code=404, detail="Campaign not found")

    stmt = (
        select(CampaignTarget)
        .where(CampaignTarget.campaign_id == campaign_id)
        .order_by(CampaignTarget.created_at)
    )
    targets = (await db.execute(stmt)).scalars().all()
    counts = await _campaign_counts(db, campaign_id)

    base = _campaign_out(campaign, counts)
    return CampaignDetailOut(
        **base.model_dump(),
        targets=[CampaignTargetOut.model_validate(t) for t in targets],
    )


@router.post("/campaigns/{campaign_id}/pause", response_model=CampaignStatusUpdateOut)
async def pause_campaign(
    campaign_id: UUID,
    db: DbDep,
    x_admin_token: str | None = Header(None),
) -> CampaignStatusUpdateOut:
    campaign = await db.get(CallCampaign, campaign_id)
    if campaign is None:
        raise HTTPException(status_code=404, detail="Campaign not found")
    campaign.status = "paused"
    await db.commit()
    return CampaignStatusUpdateOut(id=campaign.id, status=campaign.status)


@router.post("/campaigns/{campaign_id}/resume", response_model=CampaignStatusUpdateOut)
async def resume_campaign(
    campaign_id: UUID,
    db: DbDep,
    x_admin_token: str | None = Header(None),
) -> CampaignStatusUpdateOut:
    """Starts a campaign now — covers draft, scheduled, and paused campaigns
    alike, since "resume" just means "the dialer/dispatcher may act on this
    campaign starting now"."""
    campaign = await db.get(CallCampaign, campaign_id)
    if campaign is None:
        raise HTTPException(status_code=404, detail="Campaign not found")
    campaign.status = "running"
    # Manually starting a scheduled campaign early shouldn't leave a stale
    # future timestamp displayed once it's already running.
    campaign.scheduled_start_at = None
    await db.commit()
    return CampaignStatusUpdateOut(id=campaign.id, status=campaign.status)


@router.post("/campaigns/{campaign_id}/schedule", response_model=CampaignStatusUpdateOut)
async def schedule_campaign(
    campaign_id: UUID,
    payload: CampaignScheduleIn,
    db: DbDep,
    x_admin_token: str | None = Header(None),
) -> CampaignStatusUpdateOut:
    campaign = await db.get(CallCampaign, campaign_id)
    if campaign is None:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if payload.scheduled_start_at <= datetime.now(UTC):
        raise HTTPException(status_code=400, detail="scheduled_start_at must be in the future")
    campaign.status = "scheduled"
    campaign.scheduled_start_at = payload.scheduled_start_at
    await db.commit()
    return CampaignStatusUpdateOut(id=campaign.id, status=campaign.status)


@router.get("/settings")
async def get_settings(
    x_admin_token: str | None = Header(None),
) -> dict:
    return {
        "environment": settings.environment,
        "default_org_id": str(settings.default_org_id),
        "log_level": settings.log_level,
    }


@router.get("/settings/whatsapp", response_model=WhatsAppSettingsOut)
async def get_whatsapp_settings(
    x_admin_token: str | None = Header(None),
) -> WhatsAppSettingsOut:
    """Read-only status of the WhatsApp/Meta channel config for the
    /whatsapp/settings page. View-only: these are Render env vars
    (apps/api/config.py), not DB rows — a persisted override would create a
    second source of truth that can silently diverge from them (see the
    OPENAI_CHAT_MODEL incident in project notes).
    """
    return WhatsAppSettingsOut(
        configured=bool(settings.meta_access_token and settings.meta_phone_number_id),
        app_id_configured=bool(settings.meta_app_id),
        app_secret_configured=bool(settings.meta_app_secret),
        verify_token_configured=bool(settings.meta_verify_token),
        access_token_configured=bool(settings.meta_access_token),
        phone_number_id=settings.meta_phone_number_id,
        whatsapp_business_account_id=settings.meta_whatsapp_business_account_id,
        graph_api_version=settings.meta_graph_api_version,
        webhook_url=f"{settings.public_base_url.rstrip('/')}/webhook/whatsapp",
    )


@router.get("/settings/calling", response_model=CallingSettingsOut)
async def get_calling_settings(
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> CallingSettingsOut:
    """Status of the Plivo voice channel config for the /calling/settings
    page, plus the calling org's own `preferred_provider` override — see
    get_whatsapp_settings for why the credential-status fields stay
    view-only; `preferred_provider` is the one editable field here (see
    `update_calling_settings` below).
    """
    record = await db.get(Org, org)
    return CallingSettingsOut(
        configured=voice_plivo.is_configured(),
        auth_id_configured=bool(settings.plivo_auth_id),
        auth_token_configured=bool(settings.plivo_auth_token),
        phone_number=settings.plivo_phone_number,
        answer_webhook_url=f"{settings.public_base_url.rstrip('/')}/voice/answer",
        preferred_provider=record.preferred_voice_provider if record else None,
    )


@router.put("/settings/calling", response_model=CallingSettingsOut)
async def update_calling_settings(
    body: CallingSettingsIn,
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> CallingSettingsOut:
    """Set (or, with `preferred_provider: null`, clear back to automatic)
    this org's preferred voice provider — applied at every outbound-calling
    entry point via `initiate_call`'s `preferred_provider` kwarg (single
    admin call, AI callback, campaign dialer, follow-up dispatcher)."""
    record = await db.get(Org, org)
    if record is None:
        raise HTTPException(status_code=404, detail="Org not found")
    record.preferred_voice_provider = body.preferred_provider
    await db.commit()
    return CallingSettingsOut(
        configured=voice_plivo.is_configured(),
        auth_id_configured=bool(settings.plivo_auth_id),
        auth_token_configured=bool(settings.plivo_auth_token),
        phone_number=settings.plivo_phone_number,
        answer_webhook_url=f"{settings.public_base_url.rstrip('/')}/voice/answer",
        preferred_provider=record.preferred_voice_provider,
    )


# ---------------------------------------------------------------------------
# §2.6 — Day 1 control-plane endpoints
# ---------------------------------------------------------------------------


@router.get("/prompts", response_model=PromptsOut)
async def get_prompts(
    x_admin_token: str | None = Header(None),
) -> PromptsOut:
    """Read-only view of the active system prompts."""
    return PromptsOut(
        base=OUTBOUND_CALL_PROMPT,
        voice_append=VOICE_APPEND,
        whatsapp_append=WHATSAPP_APPEND,
    )


@router.get("/tools")
async def get_tools(
    x_admin_token: str | None = Header(None),
) -> list[dict]:
    """Read-only view of the registered tool schemas."""
    return TOOL_DEFINITIONS


@router.get("/script", response_model=ScriptOut)
async def get_script(
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> ScriptOut:
    """The calling org's active script — its own override if set, else the
    platform default (`OUTBOUND_CALL_PROMPT`)."""
    record = await db.get(Org, org)
    if record is not None and record.script:
        return ScriptOut(script=record.script, is_default=False)
    return ScriptOut(script=OUTBOUND_CALL_PROMPT.strip(), is_default=True)


@router.put("/script", response_model=ScriptOut)
async def update_script(
    body: ScriptIn,
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> ScriptOut:
    """Set (or, with an empty body, clear) the calling org's script override.
    Both the WhatsApp and voice channels pick this up on their next turn/call
    (see `core/agent.py::_system_prompt_for` and
    `channels/voice/realtime_bridge.py::_system_instructions`)."""
    record = await db.get(Org, org)
    if record is None:
        raise HTTPException(status_code=404, detail="Org not found")
    record.script = body.script.strip() if body.script and body.script.strip() else None
    await db.commit()
    if record.script:
        return ScriptOut(script=record.script, is_default=False)
    return ScriptOut(script=OUTBOUND_CALL_PROMPT.strip(), is_default=True)


@router.get("/org-numbers", response_model=OrgNumbersOut)
async def get_org_numbers(
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> OrgNumbersOut:
    """The calling org's own WhatsApp/calling numbers — which inbound
    messages/calls on those numbers get attributed to this org instead of
    the platform default (see channels/whatsapp/adapter.py and
    channels/voice/webhook.py)."""
    record = await db.get(Org, org)
    if record is None:
        raise HTTPException(status_code=404, detail="Org not found")
    return OrgNumbersOut(
        whatsapp_phone_number_id=record.whatsapp_phone_number_id,
        plivo_phone_number=record.plivo_phone_number,
        twilio_phone_number=record.twilio_phone_number,
    )


@router.put("/org-numbers", response_model=OrgNumbersOut)
async def update_org_numbers(
    body: OrgNumbersIn,
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> OrgNumbersOut:
    """Set (or, with an empty string, clear) this org's WhatsApp/calling
    numbers. Each field is independently optional — a field omitted from the
    request body (as opposed to sent as `null`/`""`) is left untouched, so
    the calling/WhatsApp settings pages can each save just their own number
    without clobbering the other.

    Plivo and Twilio numbers are independent fields — an org can have BOTH a
    dedicated Plivo and a dedicated Twilio number at once (the caller states
    which provider each number belongs to; this endpoint no longer guesses
    via channels/voice/number_provider.py::detect_provider or clears the
    other provider's number). Stored digits-only either way, matching how
    channels/voice/webhook.py normalizes the provider's `To` param before
    comparing.
    """
    record = await db.get(Org, org)
    if record is None:
        raise HTTPException(status_code=404, detail="Org not found")
    fields = body.model_dump(exclude_unset=True)
    if "whatsapp_phone_number_id" in fields:
        value = fields["whatsapp_phone_number_id"]
        record.whatsapp_phone_number_id = value.strip() if value else None
    if "plivo_phone_number" in fields:
        value = fields["plivo_phone_number"]
        record.plivo_phone_number = re.sub(r"\D", "", value) if value else None
    if "twilio_phone_number" in fields:
        value = fields["twilio_phone_number"]
        record.twilio_phone_number = re.sub(r"\D", "", value) if value else None
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=409, detail="One of these numbers is already assigned to another org."
        )
    return OrgNumbersOut(
        whatsapp_phone_number_id=record.whatsapp_phone_number_id,
        plivo_phone_number=record.plivo_phone_number,
        twilio_phone_number=record.twilio_phone_number,
    )


async def _claimant_names(db: DbDep, leads: Sequence[Lead]) -> dict[UUID, str]:
    """Batch-resolve claimed_by_account_user_id -> display name (full_name,
    falling back to email) for a set of leads, one query instead of N."""
    claimant_ids = {lead.claimed_by_account_user_id for lead in leads if lead.claimed_by_account_user_id}
    if not claimant_ids:
        return {}
    rows = (
        await db.execute(select(AccountUser).where(AccountUser.id.in_(claimant_ids)))
    ).scalars().all()
    return {row.id: row.full_name or row.email for row in rows}


def _lead_out_with_claimant(lead: Lead, claimant_names: dict[UUID, str]) -> LeadOut:
    out = LeadOut.model_validate(lead)
    if lead.claimed_by_account_user_id:
        out.claimed_by_name = claimant_names.get(lead.claimed_by_account_user_id)
    return out


@router.get("/escalations")
async def get_escalations(
    db: DbDep,
    redis: RedisDep,
    scope_org_id: AnalyticsScopeDep,
    x_admin_token: str | None = Header(None),
    channel: str | None = Query(None, pattern="^(voice|whatsapp)$"),
) -> dict:
    """Return recent escalation Lead rows plus the live human_handoff_queue,
    both scoped to the caller's own org unless they're the platform admin.
    The queue is a single Redis list with no per-org key, so it's filtered
    in Python on the `org_id` each entry carries (written by
    core/tools.py's transfer_to_human).
    """
    stmt = (
        select(Lead)
        .where(Lead.intent == "escalation")
        .order_by(Lead.created_at.desc())
        .limit(50)
    )
    if scope_org_id is not None:
        stmt = stmt.where(Lead.org_id == scope_org_id)
    if channel:
        stmt = stmt.where(Lead.channel == channel)
    lead_rows = (await db.execute(stmt)).scalars().all()
    claimant_names = await _claimant_names(db, lead_rows)
    recent_leads = [
        _lead_out_with_claimant(lead, claimant_names).model_dump(mode="json") for lead in lead_rows
    ]

    # LRANGE for inspection — non-destructive. Claiming (PATCH
    # /escalations/{lead_id}/claim, below) only applies to `recent_leads`
    # rows; a raw queue entry has no id to claim until transfer_to_human's
    # Lead write lands, which is the normal case since a Lead is written
    # whenever a user_id is available (see tools.py).
    raw_queue = await redis.lrange(HUMAN_HANDOFF_QUEUE, 0, -1)
    queue: list[object] = []
    for entry in raw_queue:
        # Entries are pushed as JSON by transfer_to_human; fall back to raw string
        # if a worker happens to push a plain string.
        try:
            parsed = json.loads(entry)
        except (TypeError, ValueError):
            parsed = entry
        if channel and isinstance(parsed, dict) and parsed.get("channel") != channel:
            continue
        if scope_org_id is not None:
            # A non-dict entry (plain-string fallback above) has no org to
            # check, so it can only be shown platform-wide.
            if not isinstance(parsed, dict) or parsed.get("org_id") != str(scope_org_id):
                continue
        queue.append(parsed)

    return {"recent_leads": recent_leads, "queue": queue}


@router.patch("/escalations/{lead_id}/claim", response_model=LeadOut)
async def claim_escalation(
    lead_id: UUID,
    db: DbDep,
    scope_org_id: AnalyticsScopeDep,
    current_user: CurrentUserDep,
    x_admin_token: str | None = Header(None),
) -> LeadOut:
    """A team member takes ownership of an escalation so others stop being
    alerted for it. First claim wins — a second attempt on an
    already-claimed lead is rejected (409) rather than silently
    reassigning it, so two people can't both think they own the handoff.
    """
    lead = await db.get(Lead, lead_id)
    if lead is None or lead.intent != "escalation":
        raise HTTPException(status_code=404, detail="Escalation not found")
    if scope_org_id is not None and lead.org_id != scope_org_id:
        raise HTTPException(status_code=404, detail="Escalation not found")

    if lead.claimed_by_account_user_id not in (None, current_user.id):
        claimant_names = await _claimant_names(db, [lead])
        claimant = claimant_names.get(lead.claimed_by_account_user_id, "another team member")
        raise HTTPException(status_code=409, detail=f"Already claimed by {claimant}")

    if lead.claimed_by_account_user_id is None:
        lead.claimed_by_account_user_id = current_user.id
        lead.claimed_at = datetime.now(UTC)
        await db.commit()
        await db.refresh(lead)

    return _lead_out_with_claimant(lead, {current_user.id: current_user.full_name or current_user.email})


# ---------------------------------------------------------------------------
# §3.6 — Day 2 control-plane endpoints
# ---------------------------------------------------------------------------


@router.post("/outbound/whatsapp", response_model=OutboundWhatsappOut)
@limiter.limit("30/minute")
async def outbound_whatsapp(
    request: Request,
    payload: OutboundWhatsappIn,
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> OutboundWhatsappOut:
    """Send an outbound WhatsApp message and persist the assistant turn.

    Find-or-create the user + open conversation, persist a ``Message`` row,
    then (when ``meta_access_token`` is configured) hand the body to
    ``wa_client.send_text``. When the token is unset we keep the stub
    behaviour — useful for local development without Meta credentials.
    """
    org_id = org
    usage = await get_credit_usage(db, org_id)
    await enforce_plan_limit(db, org_id, "max_whatsapp_messages", usage.whatsapp_messages)

    # Find or create the recipient user under the default org.
    user_stmt = select(User).where(User.org_id == org_id, User.phone == payload.phone)
    user = (await db.execute(user_stmt)).scalar_one_or_none()
    if user is None:
        user = User(org_id=org_id, phone=payload.phone)
        db.add(user)
        await db.flush()

    # Find an open WhatsApp conversation for this user, otherwise open a new one.
    conv_stmt = (
        select(Conversation)
        .where(
            Conversation.org_id == org_id,
            Conversation.user_id == user.id,
            Conversation.channel == "whatsapp",
            Conversation.ended_at.is_(None),
        )
        .order_by(Conversation.started_at.desc())
        .limit(1)
    )
    conversation = (await db.execute(conv_stmt)).scalar_one_or_none()
    if conversation is None:
        conversation = Conversation(org_id=org_id, user_id=user.id, channel="whatsapp")
        db.add(conversation)
        await db.flush()

    # What we record in history: the literal text, or a marker for templates.
    body_for_record = payload.text or f"[template:{payload.template_name}]"
    message = Message(
        org_id=org_id,
        conversation_id=conversation.id,
        user_id=user.id,
        role="assistant",
        content=body_for_record,
        channel="whatsapp",
    )
    db.add(message)
    await db.commit()

    # Local-dev fallback: if Meta creds are missing, keep the stub response.
    if not settings.meta_access_token:
        logger.warning(
            "outbound_whatsapp_meta_token_unset",
            phone=payload.phone,
            reason="skipping_real_send",
        )
        return OutboundWhatsappOut(status="queued", phone=payload.phone, text=body_for_record)

    # Real send. A template is the ONLY way to reach a user OUTSIDE the 24-hour
    # customer-service window (free-form text raises Meta error 131047); inside
    # the window, plain text is fine. The caller chooses by supplying
    # template_name (template) or text (free-form) — enforced by the schema.
    # Send from this org's own dedicated WhatsApp number when it has one
    # (see Org.whatsapp_phone_number_id), falling back to the platform default.
    org_record = await db.get(Org, org_id)
    phone_number_id = org_record.whatsapp_phone_number_id if org_record else None

    try:
        if payload.template_name:
            graph_response = await wa_client.send_template(
                payload.phone,
                template_name=payload.template_name,
                language_code=payload.template_lang,
                body_params=payload.template_params,
                phone_number_id=phone_number_id,
            )
        else:
            # text is guaranteed non-None here by OutboundWhatsappIn's validator.
            graph_response = await wa_client.send_text(
                payload.phone, payload.text or "", phone_number_id=phone_number_id
            )
    except httpx.HTTPError as exc:
        meta_error = wa_client._meta_error_detail(exc)
        logger.exception(
            "outbound_whatsapp_send_failed",
            phone=payload.phone,
            error=str(exc),
            meta_error=meta_error,
        )
        # Surface Meta's actual reason (unapproved/renamed template, language
        # mismatch, expired token, 24h-window violation, ...) instead of a
        # bare "WhatsApp send failed" that gives the admin nothing to act on -
        # and translate it into something actionable, not just Meta's own
        # cryptic wording.
        detail = wa_client.friendly_error_message(meta_error)
        raise HTTPException(status_code=502, detail=detail) from exc

    wa_message_id: str | None = None
    messages_block = graph_response.get("messages")
    if isinstance(messages_block, list) and messages_block:
        first = messages_block[0]
        if isinstance(first, dict):
            raw_id = first.get("id")
            if isinstance(raw_id, str):
                wa_message_id = raw_id

    logger.info(
        "outbound_whatsapp_sent",
        phone=payload.phone,
        wa_message_id=wa_message_id,
    )
    return OutboundWhatsappOut(
        status="sent",
        phone=payload.phone,
        text=body_for_record,
        wa_message_id=wa_message_id,
    )


# ---------------------------------------------------------------------------
# §4.7 — Day 3 control-plane endpoints
# ---------------------------------------------------------------------------


@router.post("/outbound/call", response_model=OutboundCallOut)
@limiter.limit("10/minute")
async def outbound_call(
    request: Request,
    payload: OutboundCallIn,
    db: DbDep,
    org: RequestOrgDep,
    x_admin_token: str | None = Header(None),
) -> OutboundCallOut:
    """Place an outbound voice call via Plivo.

    Local-dev fallback: when Plivo credentials + a from-number are not all
    configured, return a STUB response (same convention as
    ``outbound_whatsapp`` skipping the real send when ``meta_access_token`` is
    unset). The moment ``PLIVO_AUTH_ID`` / ``PLIVO_AUTH_TOKEN`` /
    ``PLIVO_PHONE_NUMBER`` are set, real calls go out automatically.

    When Plivo answers it fetches ``{PUBLIC_BASE_URL}/voice/answer`` — see
    ``channels/voice/webhook.py`` — which currently speaks a test message.
    """
    org_id = org
    usage = await get_credit_usage(db, org_id)
    await enforce_plan_limit(db, org_id, "max_call_minutes", usage.call_minutes)

    if not voice_failover.is_configured():
        logger.warning(
            "outbound_call_plivo_not_configured",
            to=payload.to_phone,
            reason="skipping_real_call",
        )
        return OutboundCallOut(call_sid=f"STUB-{uuid4()}", status="stub")

    # Dial from this org's own dedicated number(s) when it has them (see
    # Org.plivo_phone_number / Org.twilio_phone_number — independent, an org
    # can have both), falling back to each provider's platform default. Both
    # are stored digits-only (see PUT /admin/org-numbers), so re-add the "+"
    # the provider APIs expect.
    org_record = await db.get(Org, org_id)
    plivo_from = f"+{org_record.plivo_phone_number}" if org_record and org_record.plivo_phone_number else None
    twilio_from = f"+{org_record.twilio_phone_number}" if org_record and org_record.twilio_phone_number else None

    # org_id travels on the answer_url so the realtime bridge attributes this
    # call's usage to the org that actually placed it (see
    # channels/voice/webhook.py / realtime_bridge.py) rather than the
    # platform default — this is the one call site with no campaign context
    # to resolve org from instead.
    answer_url = f"{settings.public_base_url.rstrip('/')}/voice/answer?org_id={org_id}"
    # An explicit per-call choice on the dial form wins; otherwise fall back
    # to the org's persisted preference (PUT /admin/settings/calling), then
    # to failover.py's own automatic ordering if neither is set.
    preferred_provider = payload.provider or (org_record.preferred_voice_provider if org_record else None)
    try:
        result, provider = await voice_failover.initiate_call(
            payload.to_phone,
            answer_url,
            plivo_from_number=plivo_from,
            twilio_from_number=twilio_from,
            preferred_provider=preferred_provider,
        )
    except httpx.HTTPError as exc:
        logger.exception(
            "outbound_call_failed",
            to=payload.to_phone,
            error=str(exc),
        )
        raise HTTPException(status_code=502, detail="Outbound call failed") from exc

    if provider == "twilio":
        logger.warning("outbound_call_fell_back_to_twilio", to=payload.to_phone)

    request_uuid = result.get("request_uuid") or result.get("sid")
    call_sid = request_uuid if isinstance(request_uuid, str) else str(uuid4())
    return OutboundCallOut(call_sid=call_sid, status="queued")


# ---------------------------------------------------------------------------
# §5.6 — Day 4 control-plane endpoints (kill switch)
# ---------------------------------------------------------------------------


@router.post("/kill-switch", response_model=KillSwitchOut)
async def set_kill_switch(
    payload: KillSwitchIn,
    redis: RedisDep,
    x_admin_token: str | None = Header(None),
) -> KillSwitchOut:
    """Engage or release the global kill switch."""
    if payload.enabled:
        await redis.set(KILL_SWITCH_KEY, "1")
    else:
        await redis.delete(KILL_SWITCH_KEY)
    return KillSwitchOut(enabled=payload.enabled)


@router.get("/kill-switch", response_model=KillSwitchOut)
async def get_kill_switch(
    redis: RedisDep,
    x_admin_token: str | None = Header(None),
) -> KillSwitchOut:
    """Return current kill-switch state for the frontend banner."""
    value = await redis.get(KILL_SWITCH_KEY)
    return KillSwitchOut(enabled=value is not None)
