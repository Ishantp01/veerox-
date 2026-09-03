"""Self-service team management: invite/list/re-role/remove members of the
caller's own org. Unlike POST /auth/provision-org (platform-admin-only,
creates a brand new org), everything here is scoped to the org the caller's
session already belongs to via `CurrentOrgDep`.

An invite either creates a brand new AccountUser (email not seen before,
issued a fresh login token) or, if the email already has an account
elsewhere on the platform, just adds an `OrgMembership` row onto their
existing login — an AccountUser can belong to more than one org (see
`CurrentOrg` in deps.py).
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Annotated
from uuid import UUID

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select

from apps.api.core.security import generate_login_token, hash_token
from apps.api.core.sessions import invalidate_user_sessions
from apps.api.db.models.account_user import AccountUser
from apps.api.db.models.org_membership import ORG_MEMBERSHIP_ROLES, OrgMembership
from apps.api.deps import (
    CurrentOrg,
    CurrentOrgDep,
    CurrentUserDep,
    DbDep,
    RedisDep,
    enforce_plan_limit,
    require_role,
)
from apps.api.schemas.team import (
    InviteMemberIn,
    InviteMemberOut,
    RegenerateMemberTokenOut,
    TeamMemberOut,
    UpdateMemberRoleIn,
)

router = APIRouter(prefix="/team", tags=["team"])

OrgMemberDep = CurrentOrgDep
ManagerDep = Annotated[CurrentOrg, Depends(require_role("admin"))]


def _member_out(account_user: AccountUser, membership: OrgMembership) -> TeamMemberOut:
    return TeamMemberOut(
        account_user_id=account_user.id,
        email=account_user.email,
        full_name=account_user.full_name,
        mobile=account_user.mobile,
        role=membership.role,
        is_active=account_user.is_active,
        invited_at=membership.invited_at,
        joined_at=membership.joined_at,
        is_owner=membership.invited_by_id is None,
    )


@router.get("/members", response_model=list[TeamMemberOut])
async def list_members(org: OrgMemberDep, db: DbDep) -> list[TeamMemberOut]:
    result = await db.execute(
        select(AccountUser, OrgMembership)
        .join(OrgMembership, OrgMembership.account_user_id == AccountUser.id)
        .where(OrgMembership.org_id == org.org_id)
        .order_by(OrgMembership.created_at)
    )
    return [_member_out(account_user, membership) for account_user, membership in result.all()]


@router.get("/members.xlsx")
async def export_members_xlsx(org: OrgMemberDep, db: DbDep) -> StreamingResponse:
    """Same data as GET /team/members, as a downloadable .xlsx workbook."""
    result = await db.execute(
        select(AccountUser, OrgMembership)
        .join(OrgMembership, OrgMembership.account_user_id == AccountUser.id)
        .where(OrgMembership.org_id == org.org_id)
        .order_by(OrgMembership.created_at)
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Team"
    sheet.append(["name", "email", "mobile", "role", "active", "invited_at", "joined_at"])
    # Default column width is ~8.4 chars — too narrow for the formatted
    # datetime cells below (Excel shows "####" rather than truncating a
    # numeric/date value that doesn't fit, unlike a text cell) and for a
    # typical email address.
    for col, width in (("A", 20), ("B", 28), ("C", 16), ("D", 10), ("E", 8), ("F", 18), ("G", 18)):
        sheet.column_dimensions[col].width = width
    for account_user, membership in result.all():
        # The org owner (the account that bought the plan) isn't a "team
        # member" — GET /team/members' own JSON response leaves it in, but
        # the dashboard's Team page filters it out of the list it renders
        # (team/page.tsx: `!m.is_owner`); this export should match what the
        # page actually shows rather than leaking the owner's row into the
        # downloaded file.
        if membership.invited_by_id is None:
            continue
        sheet.append(
            [
                account_user.full_name or "",
                account_user.email,
                account_user.mobile or "",
                membership.role,
                account_user.is_active,
                # openpyxl only recognizes a real datetime value as a date
                # cell — a raw .isoformat() string (the old behavior) lands
                # as left-aligned text Excel won't sort/filter as a date.
                # Excel dates can't carry a timezone, so tzinfo is dropped
                # here (these are stored as UTC — see OrgMembership).
                membership.invited_at.replace(tzinfo=None) if membership.invited_at else "",
                membership.joined_at.replace(tzinfo=None) if membership.joined_at else "",
            ]
        )
        for col_idx in (6, 7):
            cell = sheet.cell(row=sheet.max_row, column=col_idx)
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="team-members.xlsx"'},
    )


@router.post("/members", response_model=InviteMemberOut, status_code=201)
async def invite_member(
    payload: InviteMemberIn,
    org: ManagerDep,
    current_user: CurrentUserDep,
    db: DbDep,
) -> InviteMemberOut:
    role = payload.role
    if role not in ORG_MEMBERSHIP_ROLES:
        raise HTTPException(status_code=400, detail=f"role must be one of {ORG_MEMBERSHIP_ROLES}")

    # The org owner (provisioned via /auth/provision-org, invited_by_id is
    # null) doesn't count against max_seats — the limit is how many
    # teammates the owner can invite, not the org's total headcount.
    invited_member_count = await db.execute(
        select(func.count())
        .select_from(OrgMembership)
        .where(OrgMembership.org_id == org.org_id, OrgMembership.invited_by_id.is_not(None))
    )
    await enforce_plan_limit(
        db,
        org.org_id,
        "max_seats",
        invited_member_count.scalar_one(),
        message="You've reached your plan's team member limit. Upgrade your plan to add more.",
    )

    existing_result = await db.execute(select(AccountUser).where(AccountUser.email == payload.email))
    account_user = existing_result.scalar_one_or_none()

    login_token: str | None = None
    if account_user is None:
        login_token = generate_login_token()
        account_user = AccountUser(
            email=payload.email,
            token_hash=hash_token(login_token),
            full_name=payload.full_name,
            mobile=payload.mobile,
        )
        db.add(account_user)
        await db.flush()
    else:
        already_member = await db.execute(
            select(OrgMembership).where(
                OrgMembership.org_id == org.org_id, OrgMembership.account_user_id == account_user.id
            )
        )
        if already_member.scalar_one_or_none() is not None:
            raise HTTPException(status_code=409, detail="This email is already on the team")
        # Existing account (already had a login elsewhere) being added to
        # this org too — let the inviter fill in a mobile it's still
        # missing, but never silently overwrite one the account already has.
        if payload.mobile and not account_user.mobile:
            account_user.mobile = payload.mobile

    db.add(
        OrgMembership(
            org_id=org.org_id,
            account_user_id=account_user.id,
            role=role,
            invited_by_id=current_user.id,
            invited_at=datetime.now(UTC),
            joined_at=datetime.now(UTC),
        )
    )
    await db.commit()

    return InviteMemberOut(
        account_user_id=account_user.id, email=account_user.email, role=role, login_token=login_token
    )


async def _get_membership(db: DbDep, org_id: UUID, account_user_id: UUID) -> OrgMembership:
    result = await db.execute(
        select(OrgMembership).where(
            OrgMembership.org_id == org_id, OrgMembership.account_user_id == account_user_id
        )
    )
    membership = result.scalar_one_or_none()
    if membership is None:
        raise HTTPException(status_code=404, detail="Team member not found")
    return membership


async def _other_admins_count(db: DbDep, org_id: UUID, excluding_account_user_id: UUID) -> int:
    result = await db.execute(
        select(func.count())
        .select_from(OrgMembership)
        .where(
            OrgMembership.org_id == org_id,
            OrgMembership.role == "admin",
            OrgMembership.account_user_id != excluding_account_user_id,
        )
    )
    return result.scalar_one()


@router.patch("/members/{account_user_id}", response_model=TeamMemberOut)
async def update_member_role(
    account_user_id: UUID, payload: UpdateMemberRoleIn, org: ManagerDep, db: DbDep
) -> TeamMemberOut:
    if payload.role not in ORG_MEMBERSHIP_ROLES:
        raise HTTPException(status_code=400, detail=f"role must be one of {ORG_MEMBERSHIP_ROLES}")

    membership = await _get_membership(db, org.org_id, account_user_id)
    if membership.role == "admin" and payload.role != "admin":
        if await _other_admins_count(db, org.org_id, account_user_id) == 0:
            raise HTTPException(status_code=400, detail="Org must keep at least one admin")

    membership.role = payload.role
    await db.commit()

    account_user_result = await db.execute(select(AccountUser).where(AccountUser.id == account_user_id))
    account_user = account_user_result.scalar_one()
    return _member_out(account_user, membership)


@router.post("/members/{account_user_id}/regenerate-token", response_model=RegenerateMemberTokenOut)
async def regenerate_member_token(
    account_user_id: UUID, org: ManagerDep, redis: RedisDep, db: DbDep
) -> RegenerateMemberTokenOut:
    """For when a teammate's login token is lost or compromised — there's no
    password reset (see core/security.py), so this is the only recovery
    path. Issues a fresh token and immediately invalidates their existing
    sessions, mirroring billing.py's regenerate_admin_token but scoped to
    the caller's own org instead of platform-admin-only.
    """
    await _get_membership(db, org.org_id, account_user_id)  # 404 if not a member of this org
    account_user_result = await db.execute(select(AccountUser).where(AccountUser.id == account_user_id))
    account_user = account_user_result.scalar_one()

    login_token = generate_login_token()
    account_user.token_hash = hash_token(login_token)
    await db.commit()
    await invalidate_user_sessions(redis, account_user_id)

    return RegenerateMemberTokenOut(
        account_user_id=account_user.id, email=account_user.email, login_token=login_token
    )


@router.delete("/members/{account_user_id}", status_code=204)
async def remove_member(account_user_id: UUID, org: ManagerDep, redis: RedisDep, db: DbDep) -> None:
    membership = await _get_membership(db, org.org_id, account_user_id)
    if membership.role == "admin" and await _other_admins_count(db, org.org_id, account_user_id) == 0:
        raise HTTPException(status_code=400, detail="Org must keep at least one admin")

    await db.delete(membership)
    await db.commit()
    await invalidate_user_sessions(redis, account_user_id)
